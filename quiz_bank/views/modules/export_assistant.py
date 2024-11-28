from django.http import HttpResponse, JsonResponse
from ...models import QuizBank, Answer
from .question_assistant import QuestionHandler, Question
from dataclasses import asdict
from openpyxl import Workbook
import tempfile
import zipfile
import os
from copy import deepcopy

OPTION_LIST = ['a', 'b', 'c', 'd', 'e', 'f', 'g']

class Filter():
    def __init__(self):
        pass
    
    def mcq_filter(self, question:Question) -> bool:
        if question.question_type == 'MCQ':
            return True
        return False

    def tf_filter(self, question:Question) -> bool:
        if question.question_type == 'TF':
            return True
        return False

    def text_filter(self, question:Question) -> bool:
        if question.question_type == 'TEXT':
            return True
        return False
    
class ExportJSON():
    def __init__(self, json_data:list[Question|dict]):
        self.json_data = [asdict(question) for question in json_data 
                          if type(question) == Question] \
                        + [question for question in json_data if type(question) == dict]

    def export(self, file_name:str):
        response = JsonResponse(self.json_data, safe=False)
        response['Content-Disposition'] = f'attachment; filename="{file_name}.json"'
        return response

class ExportPackage():
    def __init__(self):
        pass

    def __process_question_query_for_json(self, question_queryset: list[Answer]) -> list[Question]:
        question_list = QuestionHandler().process_question_query(question_queryset)
        return question_list
    
    def __question_queryset_filter(self, question_queryset: list[Answer]) -> tuple[list[Question], 
                                                                                   list[Question], 
                                                                                   list[Question]]:
        filter_obj = Filter()
        question_list = QuestionHandler().process_question_query(question_queryset)
        mcq_list = list(filter(filter_obj.mcq_filter, question_list))
        tf_list = list(filter(filter_obj.tf_filter, question_list))
        text_list = list(filter(filter_obj.text_filter, question_list))
        return mcq_list, tf_list, text_list

    def __process_question_query_for_excel(self, question_queryset: list[Answer]):
        mcq_list, tf_list, text_list = self.__question_queryset_filter(question_queryset)
        return mcq_list, tf_list, text_list

    def export_json(self, 
                    request,
                    questions_queryset:list[Answer]) -> HttpResponse:
        question_list = self.__process_question_query_for_json(questions_queryset)
        return ExportJSON(question_list).export(file_name='questions')
    
    def export_excel(self,
                     request,
                     question_queryset:list[Answer]) -> HttpResponse:
        mcq_list, tf_list, text_list = self.__process_question_query_for_excel(question_queryset)
        mcq = Workbook()
        worksheet = mcq.active
        worksheet.title = 'MCQ'
        columns = ['question', 'chapter', 'points', 'correct'] + [f'options[{option}]' for option in OPTION_LIST]
        worksheet.append(columns)
        for question in mcq_list:
            print('loop mcq')
            try:
                question_text = question.question
                chapter = question.chapter
                points = question.points
                correct = question.key
                options = question.answer
                dictionary = dict({i: OPTION_LIST[index] for index, i in enumerate(options)})
                correct_options = [dictionary[key] for key in correct]
                correct_options = ','.join(correct_options)
                
                row = [question_text, chapter, points, correct_options]
                row.extend(options)
                row = [str(col) for col in row]
                worksheet.append(row)
            except Exception as e:
                print(e)
        mcq.save('MCQ.xlsx')

        tf = Workbook()
        worksheet = tf.active
        worksheet.title = 'TF'
        columns = ['question', 'chapter', 'points', 'correct'] + [f'options[{option}]' for option in OPTION_LIST[:2]]
        worksheet.append(columns)
        for question in tf_list:
            print('loop tf')
            try:
                question_text = question.question
                chapter = question.chapter
                points = question.points
                correct = question.key
                options = question.answer

                dictionary = dict({i: OPTION_LIST[index] for index, i in enumerate(options)})
                correct_options = [dictionary[key] for key in correct]
                correct_options = ','.join(correct_options)
                
                row = [question_text, chapter, points, correct_options]
                row.extend(options)
                row = [str(col) for col in row]
                worksheet.append(row)
            except Exception as e:
                print(e)
        tf.save('TF.xlsx')

        text = Workbook()
        worksheet = text.active
        worksheet.title = 'TEXT'
        columns = ['question', 'chapter', 'points', 'correct']
        worksheet.append(columns)
        for question in text_list:
            try:
                question_text = question.question
                chapter = question.chapter
                points = question.points
                correct = question.key[0]
                
                row = [question_text, chapter, points, correct]
                row = [str(col) for col in row]
                worksheet.append(row)
            except Exception as e:
                print(e)
        text.save('TEXT.xlsx')

        # Create a temporary directory to store the files
        with tempfile.TemporaryDirectory() as tempdir:
            zip_filename = 'questions.zip'
            zip_filepath = os.path.join(tempdir, zip_filename)

            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                zip_file.write('MCQ.xlsx')
                zip_file.write('TF.xlsx')
                zip_file.write('TEXT.xlsx')

            with open(zip_filepath, 'rb') as zip_file:
                response = HttpResponse(zip_file.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="questions.zip"'
                return response
            

            
    